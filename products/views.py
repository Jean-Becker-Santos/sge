import csv
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.urls import reverse_lazy
from django.views import View
from django.http import HttpResponse
from rest_framework import generics
from openpyxl import Workbook
from products.models import Product
from products.serializers import ProductSerializer
from . import forms
from brands.models import Brand
from categories.models import Category
from app import metrics


class ProductListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Product
    template_name = 'product_list.html'
    context_object_name = 'products'
    paginate_by = 10
    permission_required = 'products.view_product'

    def get_queryset(self):
        queryset = super().get_queryset()
        title = self.request.GET.get('title')
        category = self.request.GET.get('category')
        brand = self.request.GET.get('brand')
        serie_number = self.request.GET.get('serie_number')
        if title:
            queryset = queryset.filter(title__icontains=title)
        if category:
            queryset = queryset.filter(category__id=category)
        if brand:
            queryset = queryset.filter(brand__id=brand)
        if serie_number:
            queryset = queryset.filter(brand__id=serie_number)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['product_metrics'] = metrics.get_product_metrics()
        context['categories'] = Category.objects.all()
        context['brands'] = Brand.objects.all()
        return context


class ProductCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Product
    template_name = 'product_create.html'
    form_class = forms.ProductForm
    success_url = reverse_lazy('product_list')
    permission_required = 'products.add_product'


class ProductDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Product
    template_name = 'product_detail.html'
    permission_required = 'products.view_product'


class ProductUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Product
    template_name = 'product_update.html'
    form_class = forms.ProductForm
    success_url = reverse_lazy('product_list')
    permission_required = 'products.change_product'


class ProductDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Product
    template_name = 'product_delete.html'
    success_url = reverse_lazy('product_list')
    permission_required = 'products.delete_product'


class ProductCSVExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'products.view_product'
    """
    Retorna um arquivo CSV contendo a lista de 'product',
    mas respeitando o mesmo filtro usado na productListView.
    """
    def get(self, request, *args, **kwargs):
        # 1. Capturar o parâmetro 'title' da URL (caso exista).
        title = request.GET.get('title', '')
        # 2. Iniciar o QuerySet e aplicar o filtro, se houver.
        queryset = Product.objects.all()
        if title:
            queryset = queryset.filter(title__icontains=title)
        # 3. Cria o HttpResponse com content_type de CSV e charset UTF-8
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        # 4. Define o cabeçalho para download
        response['Content-Disposition'] = 'attachment; filename="products.csv"'
        # 5. Escreve o BOM para garantir reconhecimento de UTF-8 por aplicativos como Excel
        response.write('\ufeff')  # BOM
        writer = csv.writer(response)
        # Cabeçalho do CSV
        writer.writerow(['ID', 'Produto', 'Categoria', 'Marca', 'Descrição',
                         'Número de Série', 'Preço de Custo', 'Preço de Venda',
                         'Data de Criação', 'Data de Atualização', 'Quantidade'])
        # 6. Escreve as linhas com base no QuerySet filtrado
        for product in queryset:
            writer.writerow([
                product.id,
                product.title,
                product.category,
                product.brand,
                product.description,
                product.serie_number,
                product.cost_price,
                product.selling_price,
                product.created_at,
                product.updated_at,
                product.quantity
            ])
        return response


class ProductExcelExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'products.view_product'
    """
    Retorna um arquivo Excel (XLSX) contendo a lista de 'product',
    mas respeitando o mesmo filtro usado na productListView.
    """
    def get(self, request, *args, **kwargs):
        # 1. Capturar o parâmetro 'title' da URL (caso exista).
        title = request.GET.get('title', '')
        # 2. Iniciar o QuerySet e aplicar o filtro, se houver.
        queryset = Product.objects.all()
        if title:
            queryset = queryset.filter(title__icontains=title)
        # 3. Cria uma planilha nova
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "products"
        # Define o cabeçalho
        worksheet['A1'] = 'ID'
        worksheet['B1'] = 'Produto'
        worksheet['C1'] = 'Categoria'
        worksheet['D1'] = 'Descrição'
        worksheet['E1'] = 'Número de Série'
        worksheet['F1'] = 'Preço de Custo'
        worksheet['G1'] = 'Preço de Venda'
        worksheet['H1'] = 'Data de Criação'
        worksheet['I1'] = 'Data de Atualização'
        worksheet['J1'] = 'Quantidade'
        # 4. Preenche os dados filtrados
        row_num = 2
        for product in queryset:
            worksheet.cell(row=row_num, column=1, value=product.id)
            worksheet.cell(row=row_num, column=2, value=product.title)
            worksheet.cell(row=row_num, column=3, value=str(product.category))
            worksheet.cell(row=row_num, column=4, value=product.description)
            worksheet.cell(row=row_num, column=5, value=product.serie_number)
            worksheet.cell(row=row_num, column=6, value=product.cost_price)
            worksheet.cell(row=row_num, column=7, value=product.selling_price)
            worksheet.cell(row=row_num, column=8, value=product.created_at)
            worksheet.cell(row=row_num, column=9, value=product.updated_at)
            worksheet.cell(row=row_num, column=10, value=product.quantity)
            row_num += 1
        # 5. Prepara o HttpResponse para enviar como arquivo .xlsx
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
        workbook.save(response)
        return response


class ProductCreateListAPIView(generics.ListCreateAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer


class ProductRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
