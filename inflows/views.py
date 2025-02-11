import csv
from django.views.generic import ListView, CreateView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.urls import reverse_lazy
from django.views import View
from django.http import HttpResponse
from rest_framework import generics
from openpyxl import Workbook
from inflows.models import Inflow
from inflows.serializers import InflowSerializer
from . import forms
from django.shortcuts import redirect


class InflowListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Inflow
    template_name = 'inflow_list.html'
    context_object_name = 'inflows'
    paginate_by = 10
    permission_required = 'inflows.view_inflow'

    def get(self, request, *args, **kwargs):
        page = request.GET.get('page', '')
        if '.' in page:
            # Remove todos os '.' do valor, por exemplo '151.757' => '151757'
            page_no_dot = page.replace('.', '')

            # Se por acaso ficar vazio (caso fosse só '.')
            if page_no_dot == '':
                page_no_dot = '1'
            
            # Fazemos cópia dos parâmetros GET e alteramos somente 'page'
            query_dict = request.GET.copy()
            query_dict['page'] = page_no_dot
            new_query = query_dict.urlencode()  # ex.: product=abc&page=151757
            
            # Redireciona para a mesma rota com a query "corrigida"
            # Se quiser redirecionamento permanente, use HttpResponsePermanentRedirect
            return redirect(f"{request.path}?{new_query}")

        # Caso não tenha '.', segue comportamento normal
        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        queryset = super().get_queryset()
        product = self.request.GET.get('product')
        if product:
            queryset = queryset.filter(product__title__icontains=product)
        return queryset


class InflowCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Inflow
    template_name = 'inflow_create.html'
    form_class = forms.InflowForm
    success_url = reverse_lazy('inflow_list')
    permission_required = 'inflows.view_inflow'


class InflowDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Inflow
    template_name = 'inflow_detail.html'
    permission_required = 'inflows.view_inflow'


class InflowCSVExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'inflows.view_inflow'
    """
    Retorna um arquivo CSV contendo a lista de 'Inflow',
    mas respeitando o mesmo filtro usado na InflowListView.
    """
    def get(self, request, *args, **kwargs):
        # 1. Capturar o parâmetro 'product' da URL (caso exista).
        product = request.GET.get('product', '')
        # 2. Iniciar o QuerySet e aplicar o filtro, se houver.
        queryset = Inflow.objects.all()
        if product:
            queryset = queryset.filter(product__title__icontains=product)
        # 3. Cria o HttpResponse com content_type de CSV e charset UTF-8
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        # 4. Define o cabeçalho para download
        response['Content-Disposition'] = 'attachment; filename="Inflows.csv"'
        # 5. Escreve o BOM para garantir reconhecimento de UTF-8 por aplicativos como Excel
        response.write('\ufeff')  # BOM
        writer = csv.writer(response)
        # Cabeçalho do CSV
        writer.writerow(['ID', 'Produto', 'Fornecedor', 'Data de Alteração', 'Data de Criação', 'Descrição'])
        # 6. Escreve as linhas com base no QuerySet filtrado
        for inflow in queryset:  # Variável renomeada para 'inflow'
            writer.writerow([
                inflow.id,
                inflow.product.title if inflow.product else '',  # Assegura que o produto tenha um título
                inflow.supplier.name if inflow.supplier else '',  # Assegura que o fornecedor tenha um nome
                inflow.updated_at.strftime('%d-%m-%Y %H:%M:%S') if inflow.updated_at else '',
                inflow.created_at.strftime('%d-%m-%Y %H:%M:%S') if inflow.created_at else '',
                inflow.description
            ])
        return response


class InflowExcelExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'inflows.view_inflow'
    """
    Retorna um arquivo Excel (XLSX) contendo a lista de 'Inflow',
    mas respeitando o mesmo filtro usado na InflowListView.
    """
    def get(self, request, *args, **kwargs):
        # 1. Capturar o parâmetro 'product' da URL (caso exista).
        product = request.GET.get('product', '')
        # 2. Iniciar o QuerySet e aplicar o filtro, se houver.
        queryset = Inflow.objects.all()
        if product:
            queryset = queryset.filter(product__title__icontains=product)
        # 3. Cria uma planilha nova
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Inflows"
        # Define o cabeçalho
        worksheet['A1'] = 'ID'
        worksheet['B1'] = 'Produto'
        worksheet['C1'] = 'Fornecedor'
        worksheet['D1'] = 'Quantidade'
        worksheet['E1'] = 'Alteração'
        worksheet['F1'] = 'Criação'
        worksheet['G1'] = 'Descrição'
        # 4. Preenche os dados filtrados
        row_num = 2
        for inflow in queryset:
            worksheet.cell(row=row_num, column=1, value=inflow.id)
            worksheet.cell(row=row_num, column=2, value=inflow.product.title)
            worksheet.cell(row=row_num, column=3, value=inflow.supplier.name)
            worksheet.cell(row=row_num, column=4, value=inflow.quantity)
            worksheet.cell(row=row_num, column=5, value=inflow.updated_at.strftime('%d-%m-%Y %H:%M:%S'))
            worksheet.cell(row=row_num, column=6, value=inflow.created_at.strftime('%d-%m-%Y%H:%M:%S'))
            worksheet.cell(row=row_num, column=7, value=inflow.description)
            row_num += 1
        # 5. Prepara o HttpResponse para enviar como arquivo .xlsx
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Inflows.xlsx"'
        workbook.save(response)
        return response


class InflowCreateListAPIView(generics.ListCreateAPIView):
    queryset = Inflow.objects.all()
    serializer_class = InflowSerializer


class InflowRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Inflow.objects.all()
    serializer_class = InflowSerializer
