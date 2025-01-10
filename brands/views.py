import csv
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.views import View
from django.http import HttpResponse
from openpyxl import Workbook
from brands.models import Brand
from . import forms 

class BrandListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Brand
    template_name = 'brand_list.html'
    context_object_name = 'brands'
    paginate_by = 10
    permission_required = 'brands.view_brand'

    def get_queryset(self):
        queryset = super().get_queryset()
        name = self.request.GET.get('name')

        if name:
            queryset = queryset.filter(name__icontains=name)

        return queryset

class BrandCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Brand
    template_name = 'brand_create.html' 
    form_class = forms.BrandForm
    success_url = reverse_lazy('brand_list')
    permission_required = 'brands.add_brand'


class BrandDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Brand
    template_name = 'brand_detail.html'
    permission_required = 'brands.view_brand'


class BrandUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Brand
    template_name = 'brand_update.html'
    form_class = forms.BrandForm
    success_url = reverse_lazy('brand_list')
    permission_required = 'brands.change_brand'

class BrandDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Brand
    template_name = 'brand_delete.html'
    success_url = reverse_lazy('brand_list')
    permission_required = 'brands.delete_brand'


class BrandCSVExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'brands.view_brand'
    """
    Retorna um arquivo CSV contendo a lista de 'Brand', 
    mas respeitando o mesmo filtro usado na BrandListView.
    """
    def get(self, request, *args, **kwargs):
        # 1. Capturar o parâmetro 'name' da URL (caso exista).
        name = request.GET.get('name', '')

        # 2. Iniciar o QuerySet e aplicar o filtro, se houver.
        queryset = Brand.objects.all()
        if name:
            queryset = queryset.filter(name__icontains=name)

        # 3. Cria o HttpResponse com content_type de CSV e charset UTF-8
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        # 4. Define o cabeçalho para download
        response['Content-Disposition'] = 'attachment; filename="brands.csv"'

        # 5. Escreve o BOM para garantir reconhecimento de UTF-8 por aplicativos como Excel
        response.write('\ufeff')  # BOM

        writer = csv.writer(response)
        # Cabeçalho do CSV
        writer.writerow(['ID', 'Nome', 'Descrição'])

        # 6. Escreve as linhas com base no QuerySet filtrado
        for brand in queryset:
            writer.writerow([
                brand.id,
                brand.name,
                brand.description
            ])

        return response


class BrandExcelExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'brands.view_brand'
    """
    Retorna um arquivo Excel (XLSX) contendo a lista de 'Brand', 
    mas respeitando o mesmo filtro usado na BrandListView.
    """
    def get(self, request, *args, **kwargs):
        # 1. Capturar o parâmetro 'name' da URL (caso exista).
        name = request.GET.get('name', '')

        # 2. Iniciar o QuerySet e aplicar o filtro, se houver.
        queryset = Brand.objects.all()
        if name:
            queryset = queryset.filter(name__icontains=name)

        # 3. Cria uma planilha nova
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Brands"

        # Define o cabeçalho
        worksheet['A1'] = 'ID'
        worksheet['B1'] = 'Nome'
        worksheet['C1'] = 'Descrição'

        # 4. Preenche os dados filtrados
        row_num = 2
        for brand in queryset:
            worksheet.cell(row=row_num, column=1, value=brand.id)
            worksheet.cell(row=row_num, column=2, value=brand.name)
            worksheet.cell(row=row_num, column=3, value=brand.description)
            row_num += 1

        # 5. Prepara o HttpResponse para enviar como arquivo .xlsx
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="brands.xlsx"'
        workbook.save(response)

        return response

