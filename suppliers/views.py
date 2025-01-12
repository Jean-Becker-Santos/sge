import csv
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.urls import reverse_lazy
from django.views import View
from django.http import HttpResponse
from rest_framework import generics
from openpyxl import Workbook
from suppliers.models import Supplier
from suppliers.serializers import SupplierSerializer
from . import forms


class SupplierListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Supplier
    template_name = 'supplier_list.html'
    context_object_name = 'suppliers'
    paginate_by = 10
    permission_required = 'suppliers.view_supplier'

    def get_queryset(self):
        queryset = super().get_queryset()
        name = self.request.GET.get('name')
        if name:
            queryset = queryset.filter(name__icontains=name)
        return queryset


class SupplierCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Supplier
    template_name = 'supplier_create.html'
    form_class = forms.SupplierForm
    success_url = reverse_lazy('supplier_list')
    permission_required = 'suppliers.add_supplier'


class SupplierDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Supplier
    template_name = 'supplier_detail.html'
    permission_required = 'suppliers.view_supplier'


class SupplierUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Supplier
    template_name = 'supplier_update.html'
    form_class = forms.SupplierForm
    success_url = reverse_lazy('supplier_list')
    permission_required = 'suppliers.change_supplier'


class SupplierDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Supplier
    template_name = 'supplier_delete.html'
    success_url = reverse_lazy('supplier_list')
    permission_required = 'suppliers.delete_supplier'


class SupplierCSVExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'suppliers.view_supplier'
    """
    Retorna um arquivo CSV contendo a lista de 'Supplier',
    mas respeitando o mesmo filtro usado na SupplierListView.
    """
    def get(self, request, *args, **kwargs):
        # 1. Capturar o parâmetro 'name' da URL (caso exista).
        name = request.GET.get('name', '')
        # 2. Iniciar o QuerySet e aplicar o filtro, se houver.
        queryset = Supplier.objects.all()
        if name:
            queryset = queryset.filter(name__icontains=name)
        # 3. Cria o HttpResponse com content_type de CSV e charset UTF-8
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        # 4. Define o cabeçalho para download
        response['Content-Disposition'] = 'attachment; filename="suppliers.csv"'
        # 5. Escreve o BOM para garantir reconhecimento de UTF-8 por aplicativos como Excel
        response.write('\ufeff')  # BOM
        writer = csv.writer(response)
        # Cabeçalho do CSV
        writer.writerow(['ID', 'Nome', 'Descrição'])
        # 6. Escreve as linhas com base no QuerySet filtrado
        for supplier in queryset:
            writer.writerow([
                supplier.id,
                supplier.name if supplier.name else '',
                supplier.description if supplier.description else ''
            ])
        return response


class SupplierExcelExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'suppliers.view_supplier'
    """
    Retorna um arquivo Excel (XLSX) contendo a lista de 'supplier',
    mas respeitando o mesmo filtro usado na supplierListView.
    """
    def get(self, request, *args, **kwargs):
        # 1. Capturar o parâmetro 'name' da URL (caso exista).
        name = request.GET.get('name', '')
        # 2. Iniciar o QuerySet e aplicar o filtro, se houver.
        queryset = Supplier.objects.all()
        if name:
            queryset = queryset.filter(name__icontains=name)
        # 3. Cria uma planilha nova
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "suppliers"
        # Define o cabeçalho
        worksheet['A1'] = 'ID'
        worksheet['B1'] = 'Nome'
        worksheet['C1'] = 'Descrição'
        # 4. Preenche os dados filtrados
        row_num = 2
        for supplier in queryset:
            worksheet.cell(row=row_num, column=1, value=supplier.id)
            worksheet.cell(row=row_num, column=2, value=supplier.name)
            worksheet.cell(row=row_num, column=3, value=supplier.description)
            row_num += 1
        # 5. Prepara o HttpResponse para enviar como arquivo .xlsx
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="suppliers.xlsx"'
        workbook.save(response)
        return response


class SupplierCreateListAPIView(generics.ListCreateAPIView):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer


class SupplierRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
