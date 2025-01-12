import csv
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.urls import reverse_lazy
from django.views import View
from django.http import HttpResponse
from rest_framework import generics
from openpyxl import Workbook
from categories.models import Category
from categories.serializers import CategorySerializer
from . import forms


class CategoryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Category
    template_name = 'category_list.html'
    context_object_name = 'categories'
    paginate_by = 10
    permission_required = 'categories.view_category'

    def get_queryset(self):
        queryset = super().get_queryset()
        name = self.request.GET.get('name')

        if name:
            queryset = queryset.filter(name__icontains=name)

        return queryset


class CategoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Category
    template_name = 'category_create.html'
    form_class = forms.CategoryForm
    success_url = reverse_lazy('category_list')
    permission_required = 'categories.add_category'


class CategoryDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Category
    template_name = 'category_detail.html'
    permission_required = 'categories.view_category'


class CategoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Category
    template_name = 'category_update.html'
    form_class = forms.CategoryForm
    success_url = reverse_lazy('category_list')
    permission_required = 'categories.change_category'


class CategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Category
    template_name = 'category_delete.html'
    success_url = reverse_lazy('category_list')
    permission_required = 'categories.delete_category'


class CategoryCSVExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'categories.view_category'
    """Retorna um arquivo CSV contendo a lista de 'Category',
    mas respeitando o mesmo filtro usado na CategoryListView.
    """
    def get(self, request, *args, **kwargs):
        # 1. Capturar o parâmetro 'name' da URL (caso exista).
        name = request.GET.get('name', '')
        # 2. Iniciar o QuerySet e aplicar o filtro, se houver.
        queryset = Category.objects.all()
        if name:
            queryset = queryset.filter(name__icontains=name)
        # 3. Cria o HttpResponse com content_type de CSV e charset UTF-8
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        # 4. Define o cabeçalho para download
        response['Content-Disposition'] = 'attachment; filename="categories.csv"'
        # 5. Escreve o BOM para garantir reconhecimento de UTF-8 por aplicativos como Excel
        response.write('\ufeff')  # BOM
        writer = csv.writer(response)
        # Cabeçalho do CSV
        writer.writerow(['ID', 'Nome', 'Descrição'])
        # 6. Escreve as linhas com base no QuerySet filtrado
        for category in queryset:
            writer.writerow([
                category.id,
                category.name if category.name else '',
                category.description if category.description else ''
            ])
        return response


class CategoryExcelExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'categories.view_category'
    """Retorna um arquivo Excel (XLSX) contendo a lista de 'Category',
    mas respeitando o mesmo filtro usado na CategoryListView.
    """
    def get(self, request, *args, **kwargs):
        # 1. Capturar o parâmetro 'name' da URL (caso exista).
        name = request.GET.get('name', '')
        # 2. Iniciar o QuerySet e aplicar o filtro, se houver.
        queryset = Category.objects.all()
        if name:
            queryset = queryset.filter(name__icontains=name)
        # 3. Cria uma planilha nova
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Categories"
        # Define o cabeçalho
        worksheet['A1'] = 'ID'
        worksheet['B1'] = 'Nome'
        worksheet['C1'] = 'Descrição'
        # 4. Preenche os dados filtrados
        row_num = 2
        for category in queryset:
            worksheet.cell(row=row_num, column=1, value=category.id)
            worksheet.cell(row=row_num, column=2, value=category.name)
            worksheet.cell(row=row_num, column=3, value=category.description)
            row_num += 1
        # 5. Prepara o HttpResponse para enviar como arquivo .xlsx
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="categories.xlsx"'
        workbook.save(response)
        return response


class CategoryCreateListAPIView(generics.ListCreateAPIView):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer


class CategoryRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
