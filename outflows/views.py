import csv
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.views import View
from django.http import HttpResponse
from openpyxl import Workbook
from outflows.models import Outflow
from . import forms 

class OutflowListView(ListView):
    model = Outflow
    template_name = 'outflow_list.html'
    context_object_name = 'outflows'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()
        product = self.request.GET.get('product')

        if product:
            queryset = queryset.filter(product__title__icontains=product)

        return queryset

class OutflowCreateView(CreateView):
    model = Outflow
    template_name = 'outflow_create.html' 
    form_class = forms.OutflowForm
    success_url = reverse_lazy('outflow_list')


class OutflowDetailView(DetailView):
    model = Outflow
    template_name = 'outflow_detail.html'


class OutflowCSVExportView(View):
    """
    Retorna um arquivo CSV contendo a lista de 'outflow', 
    mas respeitando o mesmo filtro usado na outflowListView.
    """
    def get(self, request, *args, **kwargs):
        # 1. Capturar o parâmetro 'product' da URL (caso exista).
        product = request.GET.get('product', '')

        # 2. Iniciar o QuerySet e aplicar o filtro, se houver.
        queryset = Outflow.objects.all()
        if product:
            queryset = queryset.filter(product__title__icontains=product)

        # 3. Cria o HttpResponse com content_type de CSV e charset UTF-8
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        # 4. Define o cabeçalho para download
        response['Content-Disposition'] = 'attachment; filename="outflows.csv"'

        # 5. Escreve o BOM para garantir reconhecimento de UTF-8 por aplicativos como Excel
        response.write('\ufeff')  # BOM

        writer = csv.writer(response)
        # Cabeçalho do CSV
        writer.writerow(['ID', 'Produto', 'Data de Alteração', 'Data de Criação', 'Descrição'])

        # 6. Escreve as linhas com base no QuerySet filtrado
        for outflow in queryset:  # Variável renomeada para 'outflow'
            writer.writerow([
                outflow.id,
                outflow.product.title if outflow.product else '',  # Assegura que o produto tenha um título
                outflow.updated_at.strftime('%d-%m-%Y %H:%M:%S') if outflow.updated_at else '',
                outflow.created_at.strftime('%d-%m-%Y %H:%M:%S') if outflow.created_at else '',
                outflow.description
            ])

        return response


class OutflowExcelExportView(View):
    """
    Retorna um arquivo Excel (XLSX) contendo a lista de 'outflow', 
    mas respeitando o mesmo filtro usado na outflowListView.
    """
    def get(self, request, *args, **kwargs):
        # 1. Capturar o parâmetro 'product' da URL (caso exista).
        product = request.GET.get('product', '')

        # 2. Iniciar o QuerySet e aplicar o filtro, se houver.
        queryset = Outflow.objects.all()
        if product:
            queryset = queryset.filter(product__title__icontains=product)

        # 3. Cria uma planilha nova
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "outflows"

        # Define o cabeçalho
        worksheet['A1'] = 'ID'
        worksheet['B1'] = 'Produto'
        worksheet['C1'] = 'Quantidade'
        worksheet['D1'] = 'Alteração'
        worksheet['E1'] = 'Criação'
        worksheet['F1'] = 'Descrição'

        # 4. Preenche os dados filtrados
        row_num = 2
        for outflow in queryset:
            worksheet.cell(row=row_num, column=1, value=outflow.id)
            worksheet.cell(row=row_num, column=2, value=outflow.product.title)
            worksheet.cell(row=row_num, column=3, value=outflow.quantity)
            worksheet.cell(row=row_num, column=4, value=outflow.updated_at.strftime('%d-%m-%Y %H:%M:%S'))
            worksheet.cell(row=row_num, column=5, value=outflow.created_at.strftime('%d-%m-%Y%H:%M:%S'))
            worksheet.cell(row=row_num, column=6, value=outflow.description)
            row_num += 1

        # 5. Prepara o HttpResponse para enviar como arquivo .xlsx
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="outflows.xlsx"'
        workbook.save(response)

        return response

